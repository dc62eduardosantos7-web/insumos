from dataclasses import dataclass, field
from decimal import Decimal
from unicodedata import normalize

from django.contrib.auth import get_user_model
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.db import transaction
from openpyxl import load_workbook

from apps.lojas.models import Loja

from .models import PerfilUsuario

Usuario = get_user_model()


@dataclass
class ResultadoImportacao:
    lojas_criadas: int = 0
    lojas_atualizadas: int = 0
    usuarios_criados: int = 0
    usuarios_existentes: int = 0
    senhas_redefinidas: int = 0
    erros: list[str] = field(default_factory=list)


def _texto(valor):
    if valor is None:
        return ""
    if isinstance(valor, Decimal):
        valor = float(valor)
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _cabecalho(valor):
    texto = normalize("NFKD", _texto(valor)).encode("ascii", "ignore").decode()
    return " ".join(texto.upper().split())


def _colunas_da_planilha(planilha):
    esperadas = {
        "CODIGO DA LOJA": "codigo",
        "NOME DA LOJA": "nome",
        "LOGIN": "login",
        "SENHA TEMPORARIA": "senha",
    }
    for numero_linha, linha in enumerate(planilha.iter_rows(values_only=True), 1):
        encontradas = {
            esperadas[cabecalho]: indice
            for indice, valor in enumerate(linha)
            if (cabecalho := _cabecalho(valor)) in esperadas
        }
        if set(encontradas) == set(esperadas.values()):
            return numero_linha, encontradas
    raise ValueError(
        "Não foi possível localizar as colunas Código da Loja, Nome da Loja, "
        "Login e Senha Temporária."
    )


def _perfil_existente(usuario):
    try:
        return usuario.perfil_insumos
    except PerfilUsuario.DoesNotExist:
        return None


def importar_logins_lojas(arquivo, redefinir_senhas=False):
    resultado = ResultadoImportacao()
    try:
        pasta = load_workbook(arquivo, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Não foi possível abrir a planilha XLSX.") from exc

    try:
        if "Credenciais" not in pasta.sheetnames:
            raise ValueError("A planilha precisa conter a aba Credenciais.")
        planilha = pasta["Credenciais"]
        linha_cabecalho, colunas = _colunas_da_planilha(planilha)

        for numero_linha, linha in enumerate(
            planilha.iter_rows(min_row=linha_cabecalho + 1, values_only=True),
            linha_cabecalho + 1,
        ):
            codigo = _texto(linha[colunas["codigo"]]).upper()
            nome = _texto(linha[colunas["nome"]])
            login = _texto(linha[colunas["login"]])
            senha = _texto(linha[colunas["senha"]])
            if not any((codigo, nome, login, senha)):
                continue

            try:
                if not codigo or not nome or not senha:
                    raise ValueError("código, nome e senha são obrigatórios")
                if login != codigo:
                    raise ValueError("o login deve ser igual ao código da loja")
                if len(codigo) > 30:
                    raise ValueError("o código possui mais de 30 caracteres")
                if len(nome) > 150:
                    raise ValueError("o nome da loja possui mais de 150 caracteres")
                if len(login) > Usuario._meta.get_field("username").max_length:
                    raise ValueError("o login é maior que o permitido")

                loja_criada = False
                loja_atualizada = False
                usuario_criado = False
                usuario_existente = False
                senha_redefinida = False
                with transaction.atomic():
                    loja, loja_criada = Loja.objects.get_or_create(
                        codigo=codigo,
                        defaults={"nome": nome},
                    )
                    if not loja_criada and loja.nome != nome:
                        loja.nome = nome
                        loja.save(update_fields=["nome", "atualizado_em"])
                        loja_atualizada = True

                    usuario = Usuario.objects.filter(username=login).first()
                    criado = usuario is None
                    if criado:
                        usuario = Usuario(username=login, first_name=nome[:150])
                        try:
                            validate_password(senha, user=usuario)
                        except ValidationError as exc:
                            raise ValueError("senha temporária inválida: " + " ".join(exc.messages))
                        usuario.set_password(senha)
                        usuario.save()
                        usuario_criado = True
                    else:
                        perfil_atual = _perfil_existente(usuario)
                        if perfil_atual and (
                            perfil_atual.papel != PerfilUsuario.LOJA
                            or perfil_atual.loja_id not in (None, loja.pk)
                        ):
                            raise ValueError(
                                "o login já pertence a outro perfil ou a outra loja"
                            )
                        usuario.first_name = nome[:150]
                        campos = ["first_name"]
                        if redefinir_senhas:
                            try:
                                validate_password(senha, user=usuario)
                            except ValidationError as exc:
                                raise ValueError(
                                    "senha temporária inválida: " + " ".join(exc.messages)
                                )
                            usuario.set_password(senha)
                            campos.append("password")
                            senha_redefinida = True
                        usuario.save(update_fields=campos)
                        usuario_existente = True

                    deve_trocar = criado or redefinir_senhas
                    perfil, perfil_criado = PerfilUsuario.objects.get_or_create(
                        usuario=usuario,
                        defaults={
                            "papel": PerfilUsuario.LOJA,
                            "loja": loja,
                            "ativo": True,
                            "deve_trocar_senha": deve_trocar,
                        },
                    )
                    if not perfil_criado:
                        perfil.papel = PerfilUsuario.LOJA
                        perfil.loja = loja
                        perfil.ativo = True
                        if deve_trocar:
                            perfil.deve_trocar_senha = True
                        perfil.save()

                resultado.lojas_criadas += int(loja_criada)
                resultado.lojas_atualizadas += int(loja_atualizada)
                resultado.usuarios_criados += int(usuario_criado)
                resultado.usuarios_existentes += int(usuario_existente)
                resultado.senhas_redefinidas += int(senha_redefinida)
            except (ValueError, ValidationError) as exc:
                resultado.erros.append(f"Linha {numero_linha} ({codigo or 'sem código'}): {exc}")
    finally:
        pasta.close()

    return resultado

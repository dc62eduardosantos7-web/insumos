from collections import OrderedDict
from decimal import Decimal, InvalidOperation
import re
import unicodedata

from django.db import transaction
from openpyxl import load_workbook

from .models import Produto
from .services import registrar_movimentacao


def _texto(valor):
    return re.sub(r"\s+", " ", str(valor or "")).strip()


def _chave(valor):
    texto = unicodedata.normalize("NFKD", _texto(valor))
    return "".join(c for c in texto if not unicodedata.combining(c)).upper()


def _decimal(valor):
    if valor in (None, ""):
        return Decimal("0")
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor)).quantize(Decimal("0.01"))

    texto = _texto(valor).replace(" ", "")
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")
    try:
        return Decimal(texto).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"TOTAL inválido: {valor}") from exc


def _localizar_cabecalhos(planilha):
    for numero_linha, linha in enumerate(
        planilha.iter_rows(min_row=1, max_row=min(planilha.max_row, 20), values_only=True),
        start=1,
    ):
        mapa = {_chave(valor): indice for indice, valor in enumerate(linha) if _texto(valor)}
        if "DESCRICAO" in mapa and "TOTAL" in mapa:
            return numero_linha, mapa
    raise ValueError("Não foi encontrada uma linha com os cabeçalhos DESCRIÇÃO e TOTAL.")


def importar_produtos_xlsx(arquivo, *, usuario=None, nome_arquivo="produtos.xlsx"):
    workbook = load_workbook(arquivo, read_only=True, data_only=True)
    planilha = workbook.active
    linha_cabecalho, colunas = _localizar_cabecalhos(planilha)

    registros = []
    ignorados = 0
    erros = []

    for numero_linha, linha in enumerate(
        planilha.iter_rows(min_row=linha_cabecalho + 1, values_only=True),
        start=linha_cabecalho + 1,
    ):
        descricao = _texto(linha[colunas["DESCRICAO"]] if colunas["DESCRICAO"] < len(linha) else "")
        if not descricao:
            ignorados += 1
            continue
        try:
            total = _decimal(linha[colunas["TOTAL"]] if colunas["TOTAL"] < len(linha) else 0)
            if total < 0:
                raise ValueError("TOTAL não pode ser negativo")
        except ValueError as exc:
            erros.append(f"Linha {numero_linha}: {exc}")
            continue

        tipo = ""
        if "TIPO" in colunas and colunas["TIPO"] < len(linha):
            tipo = _texto(linha[colunas["TIPO"]])
        observacao = ""
        if "OBS" in colunas and colunas["OBS"] < len(linha):
            observacao = _texto(linha[colunas["OBS"]])

        registros.append(
            {
                "codigo": f"INS-{numero_linha - linha_cabecalho:04d}",
                "descricao": descricao,
                "categoria": tipo,
                "total": total,
                "observacao": observacao,
                "numero_linha": numero_linha,
            }
        )

    nome_aba = planilha.title
    workbook.close()

    selecionados = OrderedDict()
    descartados = []
    for dados in registros:
        chave_descricao = _chave(dados["descricao"])
        atual = selecionados.get(chave_descricao)
        if atual is None:
            selecionados[chave_descricao] = dados
        elif dados["total"] > atual["total"]:
            descartados.append(atual)
            selecionados[chave_descricao] = dados
        else:
            descartados.append(dados)
    registros_selecionados = list(selecionados.values())

    existentes = {produto.codigo: produto for produto in Produto.objects.all()}

    excluidos_existentes = 0
    inativados_por_uso = 0
    for dados in descartados:
        produto = existentes.get(dados["codigo"])
        if produto is None or _chave(produto.nome) != _chave(dados["descricao"]):
            continue
        possui_pedido = produto.itens_pedido.exists()
        possui_movimento_manual = produto.movimentacoes.exclude(
            documento__startswith="IMPORTAÇÃO XLSX"
        ).exists()
        if not possui_pedido and not possui_movimento_manual:
            produto.movimentacoes.all().delete()
            produto.delete()
            existentes.pop(dados["codigo"], None)
            excluidos_existentes += 1
        else:
            produto.ativo = False
            produto.save(update_fields=["ativo", "atualizado_em"])
            inativados_por_uso += 1

    criados = atualizados = sem_alteracao = 0
    documento = f"IMPORTAÇÃO XLSX - {_texto(nome_arquivo)[:50]}"

    for dados in registros_selecionados:
        try:
            with transaction.atomic():
                produto = existentes.get(dados["codigo"])
                novo = produto is None
                if novo:
                    produto = Produto.objects.create(
                        codigo=dados["codigo"],
                        nome=dados["descricao"],
                        categoria=dados["categoria"][:100],
                        observacao=dados["observacao"],
                        unidade="UN",
                    )
                    existentes[dados["codigo"]] = produto
                else:
                    campos = []
                    categoria = dados["categoria"][:100]
                    observacao = dados["observacao"]
                    if produto.nome != dados["descricao"]:
                        produto.nome = dados["descricao"]
                        campos.append("nome")
                    if produto.categoria != categoria:
                        produto.categoria = categoria
                        campos.append("categoria")
                    if produto.observacao != observacao:
                        produto.observacao = observacao
                        campos.append("observacao")
                    if not produto.ativo:
                        produto.ativo = True
                        campos.append("ativo")
                    if campos:
                        produto.save(update_fields=campos + ["atualizado_em"])

                diferenca = dados["total"] - produto.estoque_atual
                if diferenca:
                    registrar_movimentacao(
                        produto=produto,
                        tipo="E" if diferenca > 0 else "S",
                        quantidade=abs(diferenca),
                        documento=documento,
                        observacao="Ajuste do saldo conforme TOTAL da planilha.",
                        usuario=usuario,
                    )

                if novo:
                    criados += 1
                elif diferenca or campos:
                    atualizados += 1
                else:
                    sem_alteracao += 1
        except Exception as exc:
            erros.append(f"Linha {dados['numero_linha']} - {dados['descricao']}: {exc}")

    return {
        "aba": nome_aba,
        "linhas_validas": len(registros),
        "produtos_importados": len(registros_selecionados),
        "descartados_menor_estoque": len(descartados),
        "excluidos_existentes": excluidos_existentes,
        "inativados_por_uso": inativados_por_uso,
        "criados": criados,
        "atualizados": atualizados,
        "sem_alteracao": sem_alteracao,
        "ignorados": ignorados,
        "erros": erros,
    }
